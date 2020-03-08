import openpyxl


class ExcelRead(object):
    def __init__(self, file, sheet_name):
        self.wb = openpyxl.load_workbook(file)
        self.ws = self.wb[sheet_name]

    def get_cell(self, row=None, column=None, cell=None):
        if cell:
            value = self.ws[cell].value
        else:
            value = self.ws.cell(row, column).value
        # # cell = self.ws[cell].value
        # cell = self.ws.cell(2,3).value
        return value

    def get_row(self, row, start_col=1, end_col=None):
        row_list = []
        for i in range(start_col, end_col + 1):
            cell = self.ws.cell(row, i).value
            row_list.append(cell)
        return row_list

    def get_range(self, start, end):
        range_list = self.ws[start:end]
        values = []
        for i in range_list:
            tem = []
            for j in i:
                v = j.value
                tem.append(v)
            values.append(tuple(tem))
        return values

    def get_data(self, start, end):
        data = self.get_range(start,end)
        title = data[0]
        test_data = data[1:]
        return title, test_data


if __name__ == '__main__':
    file = r'E:\workspaces\RPA_Control\wechart\testcases\test_data.xlsx'
    a = ExcelRead(file, 'Sheet1')
    # print(a.get_row(1, 2, 4))
    print(a.get_range(start='A1', end='C3'))
