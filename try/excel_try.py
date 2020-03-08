import openpyxl


class Excels(object):

    def __init__(self, path, sheet='Sheet1'):
        self.wb = openpyxl.load_workbook(path)
        self.ws = self.wb[sheet]
        self.rows = self.ws.max_row
        self.columns = self.ws.max_column

    def get_values(self, rows, cols):
        # todo 如不输入取最大rows、cols
        values = []
        for row in range(1, rows+1):
            row_list = []
            for col in range(1, cols+1):
                value = self.ws.cell(row, col).value
                try:
                    if isinstance(value, str):
                        value = eval(value)
                except NameError:
                    pass
                row_list.append(value)
            values.append(row_list)
        return tuple(values)

    def get_row_value(self, row):
        cols = self.columns
        rows = []
        for col in range(1, cols + 1):
            cell_value = self.ws.cell(row, col).value
            rows.append(cell_value)
        return tuple(rows)

fi = r'E:\workspaces\RPA_Control\try\Excel_try.xlsx'
a = Excels(path=fi, sheet='Sheet2')
print(a.get_values(2, 3))
# print(a.get_row_value(3))

# # 打开已有.xlsx后缀文件，不支持xlx文件
# wb = openpyxl.load_workbook(fi)
# # 获取所有sheet的名称
# # print(wb.sheetnames)
# # 访问单元格
# ws = wb['Sheet2'] # 注意sheet区分大小写
# # for i in range(2, 7):
# #     data = ws.cell(i, 1).value
# #     print(type(eval(data)))
# for row in ws.iter_rows(values_only=True):
#     print(row)
# print(ws.max_row)
# print(ws.max_column)