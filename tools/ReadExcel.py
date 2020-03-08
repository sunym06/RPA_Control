from openpyxl import load_workbook


class ReadExcel(object):

    _par = r'../data/data.xlsx'

    def __init__(self):
        self.wk = load_workbook(self._par)

    def get_value(self, sheet_name, row, col):
        sheet = self.wk[sheet_name]
        cell = sheet.cell(row, col).value
        return cell

    def get_row(self, sheet_name, row):
        rows = []
        sheet = self.wk[sheet_name]
        cols = self.wk[sheet_name].max_column
        for i in range(1, cols):
            value = sheet.cell(row, i).value
            if value is None:
                value = ''
            rows.append(value)
        return rows

    def get_column(self, sheet_name, column):
        cols = []
        sheet = self.wk[sheet_name]
        max_rows = sheet.max_row
        for i in range(1, max_rows):
            value = sheet.cell(i, column).value
            if value is None:
                value = ''
            cols.append(value)
        return cols

    def get_datas(self, sheet_name):
        data = []
        sheet = self.wk[sheet_name]
        rows = sheet.max_row
        columns = sheet.max_column
        for row in range(1, rows + 1):
            l = []
            for column in range(1, columns + 1):
                col = sheet.cell(row, column).value
                l.append(col)
            data.append(tuple(l))
        return data


if __name__ == "__main__":
    a = ReadExcel()
    # b = a.get_value('robot', 3, 5)
    c = a.get_row('robot', 2)
    d = a.get_datas('robot')
    f = a.get_column('robot', 3)
    e = a.wk['robot'].max_row
    print(c)
    print(f)

